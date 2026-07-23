"""用户、角色和字典的 Excel 导入导出服务。"""

from io import BytesIO
from types import SimpleNamespace

from fastapi import HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlmodel import select

from module_admin.dao.permission_dao import PermissionDao
from module_admin.dao.role_dao import RoleDao
from module_admin.dao.tenant_scope import require_tenant_id, tenant_clause
from module_admin.dao.user_dao import UserDao
from module_admin.entity.do.dictionary_do import DictDataDo, DictTypeDo
from module_admin.entity.do.role_do import RoleDo
from module_admin.entity.do.user_do import UserDo
from module_admin.entity.dto.role_dto import CreateRoleDto
from module_admin.entity.dto.user_dto import RegisterUserRequestByUsernameDto
from module_admin.service.password_policy import PasswordPolicyError, validate_password
from utils.fastapi_admin import FastApiAdmin


class ExcelService:
    """提供受 DTO 与租户约束的 Excel 数据交换。"""

    MAX_ROWS = 5000

    @staticmethod
    def _workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    @classmethod
    def _download(cls, filename: str, headers: list[str], rows: list[list[object]]):
        return StreamingResponse(
            BytesIO(cls._workbook_bytes(headers, rows)),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
            },
        )

    @classmethod
    async def _rows(
        cls, upload: UploadFile, required_headers: set[str]
    ) -> list[dict[str, object]]:
        content = await upload.read(20 * 1024 * 1024 + 1)
        if len(content) > 20 * 1024 * 1024:
            raise HTTPException(status_code=413, detail="Excel 文件过大")
        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheet = workbook.active
            iterator = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(iterator)]
        except (StopIteration, ValueError, TypeError) as exc:
            raise HTTPException(status_code=422, detail="Excel 文件格式无效") from exc
        missing = required_headers - set(headers)
        if missing:
            raise HTTPException(status_code=422, detail=f"缺少列: {sorted(missing)}")
        rows: list[dict[str, object]] = []
        for index, values in enumerate(iterator, start=2):
            if index > cls.MAX_ROWS + 1:
                raise HTTPException(status_code=413, detail="Excel 行数超过限制")
            if not any(value not in (None, "") for value in values):
                continue
            rows.append(
                {
                    header: values[pos] if pos < len(values) else None
                    for pos, header in enumerate(headers)
                }
            )
        workbook.close()
        return rows

    @staticmethod
    def _tenant_filter(model, request: Request):
        return tenant_clause(request, model)

    @classmethod
    async def build_export(
        cls, resource: str, mysql, tenant_id: int, actor_user_id: int
    ) -> tuple[str, list[str], list[list[object]]]:
        """在指定会话中生成导出数据，异步任务和同步接口共用此逻辑。"""
        request = SimpleNamespace(
            state=SimpleNamespace(
                mysql=mysql,
                tenant_id=tenant_id,
                user_id=actor_user_id,
            )
        )
        if resource == "users":
            result = await mysql.execute(
                select(UserDo)
                .where(cls._tenant_filter(UserDo, request))
                .order_by(UserDo.id)
            )
            field_permissions = {
                field_name: await PermissionDao.has_field_permission(
                    actor_user_id, "user", field_name, request
                )
                for field_name in ("email", "phone", "avatar")
            }
            return (
                "users.xlsx",
                [
                    "username",
                    "email",
                    "phone",
                    "avatar",
                    "nickname",
                    "status",
                    "dept_id",
                ],
                [
                    [
                        item.username,
                        item.email if field_permissions["email"] else None,
                        item.phone if field_permissions["phone"] else None,
                        item.avatar if field_permissions["avatar"] else None,
                        item.nickname,
                        item.status,
                        item.dept_id,
                    ]
                    for item in result.scalars().all()
                ],
            )
        if resource == "roles":
            result = await mysql.execute(
                select(RoleDo)
                .where(cls._tenant_filter(RoleDo, request))
                .order_by(RoleDo.id)
            )
            return (
                "roles.xlsx",
                ["name", "code", "description", "status", "data_scope"],
                [
                    [
                        item.name,
                        item.code,
                        item.description,
                        item.status,
                        item.data_scope,
                    ]
                    for item in result.scalars().all()
                ],
            )
        if resource == "dictionary":
            type_result = await mysql.execute(
                select(DictTypeDo)
                .where(cls._tenant_filter(DictTypeDo, request))
                .order_by(DictTypeDo.dict_id)
            )
            data_result = await mysql.execute(
                select(DictDataDo)
                .where(cls._tenant_filter(DictDataDo, request))
                .order_by(DictDataDo.dict_code)
            )
            rows = [
                [
                    "type",
                    item.dict_name,
                    item.dict_type,
                    item.status,
                    item.remark,
                    None,
                    None,
                    None,
                    None,
                ]
                for item in type_result.scalars().all()
            ]
            rows.extend(
                [
                    "data",
                    None,
                    item.dict_type,
                    item.status,
                    item.remark,
                    item.dict_sort,
                    item.dict_label,
                    item.dict_value,
                    item.dict_code,
                ]
                for item in data_result.scalars().all()
            )
            return (
                "dictionary.xlsx",
                [
                    "kind",
                    "dict_name",
                    "dict_type",
                    "status",
                    "remark",
                    "dict_sort",
                    "dict_label",
                    "dict_value",
                    "dict_code",
                ],
                rows,
            )
        raise ValueError("不支持的导出资源")

    @classmethod
    async def export_users(cls, request: Request):
        filename, headers, rows = await cls.build_export(
            "users",
            request.state.mysql,
            require_tenant_id(request),
            int(request.state.user_id),
        )
        return cls._download(filename, headers, rows)

    @classmethod
    async def import_users(cls, upload: UploadFile, request: Request) -> dict:
        rows = await cls._rows(upload, {"username", "password"})
        imported = 0
        errors: list[dict[str, object]] = []
        for row_number, row in enumerate(rows, start=2):
            try:
                username = str(row.get("username") or "").strip()
                password = str(row.get("password") or "")
                if not username or not password:
                    raise ValueError("username/password 不能为空")
                existing = await request.state.mysql.execute(
                    select(UserDo.id).where(
                        UserDo.username == username,
                        cls._tenant_filter(UserDo, request),
                    )
                )
                if existing.scalar_one_or_none() is not None:
                    raise ValueError("用户名已存在")
                validate_password(password, username)
                dto = RegisterUserRequestByUsernameDto(
                    username=username,
                    password=FastApiAdmin.password_hash(password),
                    email=str(row.get("email") or "") or None,
                    phone=str(row.get("phone") or "") or None,
                    nickname=str(row.get("nickname") or "") or None,
                    dept_id=(
                        int(row["dept_id"])
                        if row.get("dept_id") not in (None, "")
                        else None
                    ),
                    sex=str(row.get("sex") or "") or None,
                    post_ids=[],
                )
                await UserDao.create_user_by_username(dto, request)
                imported += 1
            except (ValueError, PasswordPolicyError) as exc:
                errors.append({"row": row_number, "message": str(exc)})
        return {"imported": imported, "failed": len(errors), "errors": errors}

    @classmethod
    async def export_roles(cls, request: Request):
        filename, headers, rows = await cls.build_export(
            "roles",
            request.state.mysql,
            require_tenant_id(request),
            int(request.state.user_id),
        )
        return cls._download(filename, headers, rows)

    @classmethod
    async def import_roles(cls, upload: UploadFile, request: Request) -> dict:
        rows = await cls._rows(upload, {"name", "code"})
        imported = 0
        errors: list[dict[str, object]] = []
        for row_number, row in enumerate(rows, start=2):
            try:
                data = CreateRoleDto(
                    name=str(row.get("name") or "").strip(),
                    code=str(row.get("code") or "").strip(),
                    description=str(row.get("description") or ""),
                    data_scope=str(row.get("data_scope") or "5"),
                )
                await RoleDao.create_role_by_role_name(data, request)
                imported += 1
            except (ValueError, TypeError) as exc:
                errors.append({"row": row_number, "message": str(exc)})
        return {"imported": imported, "failed": len(errors), "errors": errors}

    @classmethod
    async def export_dictionary(cls, request: Request):
        filename, headers, rows = await cls.build_export(
            "dictionary",
            request.state.mysql,
            require_tenant_id(request),
            int(request.state.user_id),
        )
        return cls._download(filename, headers, rows)

    @classmethod
    async def import_dictionary(cls, upload: UploadFile, request: Request) -> dict:
        rows = await cls._rows(upload, {"kind", "dict_type"})
        imported = 0
        errors: list[dict[str, object]] = []
        for row_number, row in enumerate(rows, start=2):
            try:
                kind = str(row.get("kind") or "").strip().lower()
                dict_type = str(row.get("dict_type") or "").strip()
                if kind == "type":
                    request.state.mysql.add(
                        DictTypeDo(
                            dict_name=str(row.get("dict_name") or "").strip(),
                            dict_type=dict_type,
                            status=str(row.get("status") or "1"),
                            remark=str(row.get("remark") or "") or None,
                            tenant_id=require_tenant_id(request),
                        )
                    )
                elif kind == "data":
                    parent = await request.state.mysql.execute(
                        select(DictTypeDo.dict_id).where(
                            DictTypeDo.dict_type == dict_type,
                            cls._tenant_filter(DictTypeDo, request),
                        )
                    )
                    if parent.scalar_one_or_none() is None:
                        raise ValueError("字典类型不存在")
                    request.state.mysql.add(
                        DictDataDo(
                            dict_sort=int(row.get("dict_sort") or 0),
                            dict_label=str(row.get("dict_label") or "").strip(),
                            dict_value=str(row.get("dict_value") or "").strip(),
                            dict_type=dict_type,
                            status=str(row.get("status") or "1"),
                            remark=str(row.get("remark") or "") or None,
                            tenant_id=require_tenant_id(request),
                        )
                    )
                else:
                    raise ValueError("kind 必须为 type 或 data")
                imported += 1
            except (ValueError, TypeError) as exc:
                errors.append({"row": row_number, "message": str(exc)})
        return {"imported": imported, "failed": len(errors), "errors": errors}
